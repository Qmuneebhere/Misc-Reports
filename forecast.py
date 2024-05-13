from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re
from datetime import datetime, date
import glob
import os
import time
import shutil

section1 = '\n\n***********************************************************\n\n'

##########################################################################
#########################                        #########################
######################### Author: Muneeb Qureshi #########################
#########################                        #########################
##########################################################################


##########################################################################
##                                                                      ##
##        This function takes folder path and file name as input        ##
##        Reads xlsx file, cleans data, performs QA on the files        ##
##        and creates a text file of the data (without headers).        ##
##                                                                      ##
##########################################################################

def ReadFile(currPath, name, dateToday):
    
    os.chdir(currPath)

    # Name of sheet in workbook File

    sheetName = 'RCS_Forecast'
    
    # Creates a workbook & worksheet object

    workbook = load_workbook(name)
    worksheet = workbook[sheetName]
    
    data = []
    
    
    #####################################################################
    ##                                                                 ##
    ##                       Condition 1 & 2:                          ##
    ##                                                                 ##
    ##       In forecast worksheet First four rows are headers         ##
    ##       'C' column is blank in headers and 'FCT' value in         ## 
    ##       other rows. So either condition1 would be true or         ##
    ##       condition2. If condition2 is true row is required         ##
    ##        data, which is then appended in list named data.         ##
    ##                                                                 ##
    #####################################################################

    
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):

        condition1 = index <= 3 and row[2] is None
        condition2 = index > 3 and row[2] == 'FCT'

        if condition1 or condition2:

            if condition2:

                data.append(list(row))

        else:

            print(section1)
            print('File is irregular, check first 4 rows and C column')
            
            
    #####################################################################
    ##                                                                 ##
    ##                         Condition 3:                            ##
    ##                                                                 ##
    ##        Condition 3 is for first two columns of dataframe        ##
    ##        If first two columns are completely empty they're        ##
    ##                          deleted.                               ##
    ##                                                                 ##
    #####################################################################

    
    df = pd.DataFrame(data, dtype='object')

    condition3 = df[0].isnull().all() and df[1].isnull().all()

    if condition3:

        df = df.drop(df.columns[:2], axis=1)

    else:

        print(section1)
        print('File is irregular, check A, B columns.')

    # Number of columns must be 16. 
    
    if df.shape[1] != 16:

        print(section1)
        print('File is irregular. Check number of Columns.')
    
    # Columns are named using following list.

    newHeaders = ['FCT', 'code', 'date', 'numeric1', 'numeric2',
                  'numeric3', 'numeric4', 'numeric5', 'empty1',
                  'numeric6', 'empty2', 'empty3', 'text1',
                  'text2', 'final1', 'final2']

    df.columns = newHeaders
    
    
    #####################################################################
    ##                                                                 ##
    ##                         Condition 4:                            ##
    ##                                                                 ##
    ##         Condition 4 is for empty columns of dataframe if        ##
    ##          all three columns are completely empty they're         ##
    ##                          deleted.                               ##
    ##                                                                 ##
    #####################################################################

    
    condition4 = df['empty1'].isnull().all() and \
                 df['empty2'].isnull().all() and \
                 df['empty3'].isnull().all()

    if condition4:

        df = df.drop(['empty1', 'empty2', 'empty3'], axis=1)

    else:

        print(section1)
        print('File is irregular, check empty merged columns.')

    numCols = ['numeric1', 'numeric2', 'numeric3',
               'numeric4', 'numeric5', 'numeric6']
    
    # Converts numeric Columns to floats
    
    df[numCols] = df[numCols].astype(float)
    
    
    #####################################################################
    ##                                                                 ##
    ##       All the elements in numeric columns must be positive      ##
    ##       and must not be blank, if there are any negatives or      ##
    ##               blanks, they must be replaced with 0.             ##
    ##                                                                 ##
    #####################################################################

    
    df[numCols] = \
    df[numCols].map(lambda x: 0.00 if (pd.isnull(x) or x < 0) else x)

    # Formats the numeric columns to numbers with 2 decimal places.

    df[numCols] = df[numCols].apply(lambda x: x.round(2))
    
    
    #####################################################################
    ##                                                                 ##
    ##                           Date Column:                          ##
    ##                                                                 ##
    ##           Converting datetime column to string column           ## 
    ##               date formatting would be m/d/yyyy                 ##
    ##                                                                 ##
    #####################################################################

    
    df['date'] = df['date'].apply(lambda x: x.strftime('%#m/%#d/%Y'))

    # Checking if column contains today's date only

    isDateValid = (df['date'] == dateToday).all()

    if not isDateValid:

        print(section1)
        print("The date in file is not matching with today's date.")
    
    # Gets date of the file and check if its unique
    
    uniqueDate = df['date'].unique()
    
    if len(uniqueDate) != 1:

        print(section1)
        print('File contains multiple dates. Check Date column.')
        
        return False
    
    else: date = uniqueDate[0]
   

    #####################################################################
    ##                                                                 ##
    ##                                QA:                              ##
    ##                                                                 ##
    ##               Checking code and numeric Columns.                ## 
    ##          1. The code column must only have '008357'             ##
    ##          2. The sum of 2nd last two columns must be             ##
    ##                   equal to sum of last column                   ##
    ##                                                                 ##
    #####################################################################
    
    
    isCodeValid = (df['code'] == '008357').all()
    
    if not isCodeValid:

        print(section1)  
        print('File is irregular, check 008357 code column.')
    
    isNumValid = (abs(df['numeric4'] + df['numeric5'] \
                  - df['numeric6']).round(2) <= 0.01).all()
    
    if isNumValid:
        
        df['numeric6'] = df['numeric4'] + df['numeric5']
    
    if not isNumValid:

        print(section1)
        print('The sum of numeric Columns is not equal' \
         'to last numeric Column.')
        
        
    #####################################################################
    ##                                                                 ##
    ##                        Save as Text File:                       ##
    ##                                                                 ##
    ##         The numeric column must be converted to string          ## 
    ##         with 2 decimal formatting before saving as text         ##
    ##         file. Because as floats, zeros are skipped. Its         ##
    ##         saved with name RCS_FCT.txt as a temporary file         ##
    ##                 which does not include headers.                 ##
    ##                                                                 ##
    #####################################################################


    df[numCols] = df[numCols].map(lambda x: "{:.2f}".format(x))
    
    rowCount = df.shape[0]

    # Saves current forecast as text file

    df.to_csv(currPath + 'RCS_FCT.txt',
              header=False, index=False, sep='\t')
    
    return date, rowCount


##########################################################################
##                                                                      ##
##  Takes date in file and formats date for both headers and filename.  ## 
##        Returns the header date and Filename (for saving file)        ##
##                                                                      ##
##########################################################################

def FileDetails(origDate):
    
    month, date, year = origDate.split('/')

    month = month.zfill(2)
    date = date.zfill(2)

    headerDate = f'{month}/{date}/{year}'
    fileDate = f'{year}{month}{date}'
    
    # File name for saving forecast file

    fctName = f'FCT_008357_01_{fileDate}_103000.txt'
    
    return headerDate, fctName


##########################################################################
##                                                                      ##
##    This function takes original name of text file containing data    ##
##    and the new name assigned to final file after including header    ##
##         It takes date in the file as input and also rowcount         ##
##         Creates a text file with new name including headers.         ##
##             Saves that text file in current date folder.             ##
##                                                                      ##
##########################################################################

def AddHeaders(currPath, newName, date, count):
    
    # Opens the original file (without headers) in read mode
    
    txtFile = currPath + 'RCS_FCT.txt'
    
    with open(txtFile, 'r') as origFile:
    
        # Read the existing content
    
        existingContent = origFile.read()
    
    # Creates a new forecast text file with headers

    h1 = f'FHD\t01\t02\tFCT\t008357\t000075\t{date} 10.30.00\tUNIFIN\n'
    h2 = f'RHD 01\tFCT 02\t{count}\n'
        
    with open(currPath + newName, 'w') as newFile:
        
        newFile.write(h1)
        newFile.write(h2)
        
        newFile.write(existingContent)


# -----------------------------Today's Date----------------------------- #


# Gets today's date as string with formatting mm/dd/yyyy

dateToday = date.today().strftime('%#m/%#d/%Y')

# Current year as string

year = date.today().strftime('%Y')

# Date as string in format mm-dd-yyyy

dateFolder = date.today().strftime('%m-%d-%Y')


# --------------Paths of folders required for saving files-------------- #


# Path to current working directory

currDir = f'{os.getcwd()}\\'


# ------------------------Get Workbook File name------------------------ #


# Changes directory to current date folder

os.chdir(currDir)

# Pattern for matching workbook file name

pattern = '*RCS_FCT*.xlsx'

# Gets List of files containing RCS_FCT in their name

fileNames = glob.glob(pattern)

# Gets the file name of todays file

fileName = fileNames[0]


# -------------------Gets date in file and rows Count------------------- #


# This function also saves temporary RCS_FCT.txt file without headers

origDate, rowCount = ReadFile(currDir, fileName, dateToday)


# ---------------Gets date and filename for forecast file--------------- #

    
headerDate, fctName = FileDetails(origDate)


# ---------Adds Header and save forecast file in current folder--------- #

    
AddHeaders(currDir, fctName, headerDate, rowCount)


# ----------Move final forecast file to respective year folder---------- #


print(section1)
print('After Checking the files press (Y): '.center(60) + '\n')

while True:
    
    inp = input('Move Forecast File: ')
    
    if inp == 'Y': break
    else: print('\nInvalid Input\n')

# Path to current year folder

yearFolder = f'{currDir}\\{year}\\'

# Creates year folder if does not exist

if not os.path.exists(yearFolder): os.makedirs(yearFolder)

# Moves file to current year folder

shutil.move(currDir + fctName, yearFolder + fctName)

print(section1)
print(f'File has been moved to {year} Folder.'.center(60))


# ------------Deletes current date folder after moving files------------ #


os.remove(currDir + 'RCS_FCT.txt')

try: os.remove(currDir + fileName)

except PermissionError:

    print(f'Close {fctName} File and press (Y):')

    while True:

        inp = input('Delete file: ')

        if inp == 'Y': break
        else: print('\nInvalid Input\n')



